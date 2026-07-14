import os
import gc
import io
import csv  # Built-in Python library (0MB RAM footprint)
from flask import Flask, request, Response
from flask_cors import CORS  
import openpyxl
from pypdf import PdfReader

try:
    import pytesseract
    from PIL import Image as PILImage
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

app = Flask(__name__)

# --- CONFIGURE CORS POLICY FOR SALESFORCE ORG ---
CORS(app, resources={
    r"/*": {
        "origins": ["https://sidd-idp2-dev-ed.trailblaze.lightning.force.com"],
        "methods": ["POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "X-File-Name"]
    }
})

@app.route('/convert', methods=['POST', 'OPTIONS'])
def convert():
    if request.method == 'OPTIONS':
        return '', 200

    file_bytes = None
    try:
        filename = request.headers.get('X-File-Name', 'document.pdf').lower()
        print(f"Processing inbound request for file: {filename}")

        # Stream the body directly into an in-memory byte buffer
        file_bytes = io.BytesIO(request.get_data(parse_form_data=False))
        
        if file_bytes.getbuffer().nbytes == 0:
            return "Empty request body", 400

        # --- ROUTE A: LIGHTWEIGHT PDF PARSING ---
        if filename.endswith('.pdf'):
            print("Routing PDF through lightweight pypdf engine...")
            reader = PdfReader(file_bytes)
            extracted_text = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    extracted_text.append(f"## Page {i+1}\n\n{text}")
            markdown_output = "\n\n".join(extracted_text)
            if not markdown_output.strip():
                markdown_output = "*No machine-readable text found in PDF.*"
            return Response(markdown_output, mimetype='text/plain')

        # --- ROUTE B: LIGHTWEIGHT EXCEL PARSING (.xlsx) ---
        elif filename.endswith(('.xlsx', '.xls')):
            print("Routing Excel through lightweight openpyxl engine...")
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            markdown_sheets = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                markdown_sheets.append(f"## Sheet: {sheet_name}\n")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_vals = [str(cell) if cell is not None else "" for cell in row]
                        markdown_sheets.append("| " + " | ".join(row_vals) + " |")
                markdown_sheets.append("\n")
            markdown_output = "\n".join(markdown_sheets)
            return Response(markdown_output, mimetype='text/plain')

        # --- NEW ROUTE C: LIGHTWEIGHT NATIVE CSV PARSING ---
        elif filename.endswith('.csv'):
            print("Routing CSV through native string memory parser...")
            # Decode binary stream into standard text characters safely
            csv_text = file_bytes.getvalue().decode('utf-8', errors='ignore')
            
            # Read split lines utilizing Python's structured string CSV parser
            csv_reader = csv.reader(io.StringIO(csv_text))
            markdown_rows = []
            
            for index, row in enumerate(csv_reader):
                if not row or all(cell.strip() == "" for cell in row):
                    continue # Skip blank spacer rows
                
                # Format elements into crisp markdown row strings
                clean_row = "| " + " | ".join([cell.replace("|", "\\|") for cell in row]) + " |"
                markdown_rows.append(clean_row)
                
                # Automatically construct the essential Markdown table separation row right below headers
                if index == 0:
                    separator = "| " + " | ".join(["---"] * len(row)) + " |"
                    markdown_rows.append(separator)
            
            markdown_output = "\n".join(markdown_rows)
            return Response(markdown_output, mimetype='text/plain')

        # --- ROUTE D: IMAGE FILES (TESSERACT) ---
        elif filename.endswith(('.png', '.jpg', '.jpeg')):
            if TESSERACT_AVAILABLE:
                print("Routing image through Tesseract engine...")
                img = PILImage.open(file_bytes)
                extracted_text = pytesseract.image_to_string(img)
                del img
                response_text = extracted_text if extracted_text.strip() else "OCR complete: No readable text found."
                return Response(response_text, mimetype='text/plain')
            else:
                return "Error: Tesseract not available inside Docker container.", 500

        # --- ROUTE E: STRUCTURAL TEXT FILES ---
        elif filename.endswith(('.xml', '.json', '.txt')):
            text_content = file_bytes.getvalue().decode('utf-8', errors='ignore')
            return Response(f"```{filename.split('.')[-1]}\n{text_content}\n```", mimetype='text/plain')

        else:
            return f"Unsupported file extension: {filename}", 400

    except Exception as e:
        print(f"!!! CRITICAL PYTHON CRASH !!!: {str(e)}")
        return f"Internal Server Error: {str(e)}", 500

    finally:
        if file_bytes:
            file_bytes.close()
        gc.collect()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
